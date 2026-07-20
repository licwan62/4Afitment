"""Shared YAML configuration and tabular input helpers for scraper projects."""

from __future__ import annotations

import csv
import json
import time
from pathlib import Path
from typing import Any


TABLE_EXTENSIONS = (".tsv", ".csv", ".xlsx", ".xlsm")
DELIMITED_TABLE_EXTENSIONS = (".tsv", ".csv")
PATH_KEYS = {
    "input",
    "output",
    "log",
    "checkpoint",
    "profile_dir",
}


def load_yaml_config(config_path: str | Path, section: str) -> dict[str, Any]:
    path = Path(config_path).resolve()
    if not path.exists():
        return {}
    try:
        import yaml
    except ImportError as exc:
        raise RuntimeError("读取 YAML 配置需要安装 PyYAML") from exc

    raw = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    if not isinstance(raw, dict):
        raise RuntimeError(f"YAML 顶层必须是对象：{path}")
    common = raw.get("common", {}) or {}
    project = raw.get(section, {}) or {}
    if not isinstance(common, dict) or not isinstance(project, dict):
        raise RuntimeError(f"YAML common/{section} 必须是对象：{path}")
    result = {**common, **project}
    for key in PATH_KEYS:
        value = result.get(key)
        if value and not Path(str(value)).is_absolute():
            result[key] = str((path.parent / str(value)).resolve())
    return result


def apply_known_defaults(parser: Any, defaults: dict[str, Any]) -> None:
    known = {action.dest for action in parser._actions}
    parser.set_defaults(**{key: value for key, value in defaults.items() if key in known})


def discover_input_file(value: str | Path) -> Path:
    path = Path(value).resolve()
    if path.is_file():
        return path
    if not path.exists():
        # CSV and TSV are interchangeable inputs throughout the scrapers.  When
        # configuration still names the old format, accept a same-stem sibling
        # written in the other format (for example catalog.csv -> catalog.tsv).
        if path.suffix.casefold() in DELIMITED_TABLE_EXTENSIONS:
            for suffix in DELIMITED_TABLE_EXTENSIONS:
                if suffix == path.suffix.casefold():
                    continue
                alternative = path.with_suffix(suffix)
                if alternative.is_file():
                    return alternative
        raise FileNotFoundError(f"输入不存在：{path}")
    if not path.is_dir():
        raise RuntimeError(f"输入必须是表格文件或目录：{path}")
    candidates = sorted(
        item for item in path.iterdir()
        if item.is_file() and item.suffix.casefold() in TABLE_EXTENSIONS
    )
    if not candidates:
        raise RuntimeError(f"输入目录中没有 TSV/CSV/XLSX 文件：{path}")
    return candidates[0]


def read_table_records(
    value: str | Path, sheetname: str | None = None
) -> tuple[Path, list[dict[str, str]]]:
    path = discover_input_file(value)
    suffix = path.suffix.casefold()
    if suffix in {".xlsx", ".xlsm"}:
        return path, _read_xlsx(path, sheetname)
    delimiter = "\t" if suffix == ".tsv" else ","
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle, delimiter=delimiter)
        if not reader.fieldnames:
            raise RuntimeError(f"输入表没有表头：{path}")
        rows = [
            {str(key): _cell_text(value) for key, value in row.items() if key is not None}
            for row in reader
        ]
    return path, rows


def _read_xlsx(path: Path, sheetname: str | None) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("读取 XLSX 需要安装 openpyxl") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheetname:
            if sheetname not in workbook.sheetnames:
                raise RuntimeError(
                    f"XLSX 中不存在 sheetname={sheetname}；可用：{', '.join(workbook.sheetnames)}"
                )
            sheet = workbook[sheetname]
        else:
            sheet = workbook[workbook.sheetnames[0]]
        iterator = sheet.iter_rows(values_only=True)
        header_row = next(iterator, None)
        if header_row is None:
            raise RuntimeError(f"XLSX sheet 为空：{path} / {sheet.title}")
        headers = [_cell_text(value) for value in header_row]
        if not any(headers):
            raise RuntimeError(f"XLSX sheet 没有表头：{path} / {sheet.title}")
        result: list[dict[str, str]] = []
        for values in iterator:
            row = {
                header: _cell_text(values[index] if index < len(values) else "")
                for index, header in enumerate(headers)
                if header
            }
            if any(row.values()):
                result.append(row)
        return result
    finally:
        workbook.close()


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return " ".join(str(value).replace("\xa0", " ").split())


def output_file(output_dir: str | Path, filename: str) -> Path:
    directory = Path(output_dir).resolve()
    directory.mkdir(parents=True, exist_ok=True)
    return directory / filename


def append_json_log(
    log_path: str | Path, level: str, event: str, detail: object
) -> None:
    path = Path(log_path).resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    entry = {
        "time": time.strftime("%Y-%m-%d %H:%M:%S"),
        "level": level,
        "event": event,
        "detail": detail,
    }
    with path.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(entry, ensure_ascii=False) + "\n")
